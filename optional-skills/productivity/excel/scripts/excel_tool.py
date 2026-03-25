#!/usr/bin/env python3
"""
Excel/XLSX Tool for Hermes Agent
----------------------------------
Read, create, edit, search, and convert Excel spreadsheets.

Requires: pip install openpyxl

Usage:
  python3 excel_tool.py info    <file>
  python3 excel_tool.py read    <file> [--sheet NAME] [--range A1:C10] [--json]
  python3 excel_tool.py create  <file> [--sheet NAME] [--data '[[...]]']
  python3 excel_tool.py edit    <file> --cell A1 --value "hello" [--sheet NAME]
  python3 excel_tool.py search  <file> <query> [--sheet NAME]
  python3 excel_tool.py to-csv  <file> [--sheet NAME] [--output out.csv]
  python3 excel_tool.py from-csv <csv_file> <xlsx_file> [--sheet NAME]
  python3 excel_tool.py add-row <file> --data '["a","b","c"]' [--sheet NAME]
"""

import argparse
import csv
import json
import os
import sys
from typing import Any, Dict, List, Optional


def _check_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        sys.exit(
            "openpyxl not installed.\n"
            "  pip install openpyxl\n"
        )


def _load_workbook(path: str, read_only: bool = False):
    import openpyxl
    if not os.path.exists(path):
        sys.exit(f"File not found: {path}")
    try:
        return openpyxl.load_workbook(path, read_only=read_only, data_only=True)
    except Exception as e:
        sys.exit(f"Failed to open {path}: {e}")


def _get_sheet(wb, sheet_name: Optional[str] = None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        return wb[sheet_name]
    return wb.active


def _cell_value(cell) -> Any:
    """Convert cell value to JSON-serializable type."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    # datetime, date, time
    try:
        return str(v)
    except Exception:
        return repr(v)


# Formula-injection prefixes — when a value starts with these characters,
# Excel interprets it as a formula.  We escape them with a leading single
# quote to prevent execution when the spreadsheet is opened.
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_value(value: Any) -> Any:
    """Escape formula-injection prefixes in string values."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value  # Excel treats leading ' as text prefix
    return value


def _safe_csv_value(value: Any) -> Any:
    """Escape formula-injection prefixes for CSV output."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_info(args):
    """Show workbook metadata: sheets, dimensions, cell counts."""
    _check_openpyxl()
    wb = _load_workbook(args.file, read_only=False)

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "dimensions": getattr(ws, "dimensions", "unknown") or "empty",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })

    out = {
        "file": args.file,
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets,
        "active_sheet": wb.active.title if wb.active else None,
    }
    wb.close()
    print(json.dumps(out, indent=2, default=str))


def cmd_read(args):
    """Read cell data from a sheet, optionally within a range."""
    _check_openpyxl()
    wb = _load_workbook(args.file, read_only=True)
    ws = _get_sheet(wb, args.sheet)

    if args.range:
        try:
            rows = []
            for row in ws[args.range]:
                rows.append([_cell_value(c) for c in row])
                if len(rows) >= args.limit:
                    break
        except Exception as e:
            sys.exit(f"Invalid range '{args.range}': {e}")
    else:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, args.limit),
                                 max_col=ws.max_column):
            rows.append([_cell_value(c) for c in row])

    out = {
        "file": args.file,
        "sheet": ws.title,
        "rows": len(rows),
        "columns": len(rows[0]) if rows else 0,
    }

    if args.headers and rows:
        headers = [str(v) if v else f"col_{i}" for i, v in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            data.append(dict(zip(headers, row)))
        out["headers"] = headers
        out["data"] = data
    else:
        out["data"] = rows

    wb.close()
    print(json.dumps(out, indent=2, default=str))


def cmd_create(args):
    """Create a new workbook with optional initial data or multi-sheet JSON structure."""
    _check_openpyxl()
    import openpyxl

    if os.path.exists(args.file) and not args.force:
        sys.exit(f"File already exists: {args.file}. Use --force to overwrite.")

    wb = openpyxl.Workbook()

    # Multi-sheet structured creation via --from-json
    if args.from_json:
        try:
            sheets_def = json.loads(args.from_json)
            if not isinstance(sheets_def, list):
                sys.exit("--from-json must be a JSON array of sheet definitions")
        except json.JSONDecodeError as e:
            sys.exit(f"Invalid JSON in --from-json: {e}")

        default_ws = wb.active
        sheets_created = []
        for i, sheet_def in enumerate(sheets_def):
            if not isinstance(sheet_def, dict):
                continue
            name = str(sheet_def.get("name", f"Sheet{i+1}"))[:31]
            data = sheet_def.get("data", [])
            style = sheet_def.get("style", True)  # header styling on by default

            if i == 0:
                default_ws.title = name
                ws = default_ws
            else:
                ws = wb.create_sheet(title=name)

            for row in data:
                if isinstance(row, list):
                    ws.append([_safe_value(v) for v in row])
                else:
                    ws.append([_safe_value(row)])

            if style and ws.max_row and ws.max_row > 0:
                _style_header(ws)
                _auto_width(ws)

            sheets_created.append({"name": name, "rows": ws.max_row or 0})

        wb.save(args.file)
        wb.close()
        print(json.dumps({
            "created": args.file,
            "sheets": sheets_created,
        }, indent=2))
        return

    # Simple single-sheet creation
    ws = wb.active
    ws.title = args.sheet or "Sheet1"

    if args.data:
        try:
            data = json.loads(args.data)
            if not isinstance(data, list):
                sys.exit("--data must be a JSON array of arrays: [[1,2],[3,4]]")
            for row in data:
                if isinstance(row, list):
                    ws.append([_safe_value(v) for v in row])
                else:
                    ws.append([_safe_value(row)])
        except json.JSONDecodeError as e:
            sys.exit(f"Invalid JSON in --data: {e}")

    wb.save(args.file)
    wb.close()
    print(json.dumps({
        "created": args.file,
        "sheet": ws.title,
        "rows": ws.max_row,
    }, indent=2))


def cmd_edit(args):
    """Edit a specific cell value."""
    _check_openpyxl()
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    old_value = ws[args.cell].value

    # Try to parse value as number
    value = args.value
    try:
        if "." in value:
            value = float(value)
        else:
            value = int(value)
    except (ValueError, TypeError):
        value = _safe_value(value)  # escape formula prefixes

    ws[args.cell] = value
    wb.save(args.file)
    wb.close()

    print(json.dumps({
        "file": args.file,
        "sheet": ws.title,
        "cell": args.cell,
        "old_value": _cell_value(type("obj", (), {"value": old_value})()),
        "new_value": value,
    }, indent=2, default=str))


def cmd_search(args):
    """Search for a value across all cells in a sheet."""
    _check_openpyxl()
    wb = _load_workbook(args.file, read_only=True)
    ws = _get_sheet(wb, args.sheet)

    query = args.query.lower()
    matches = []

    # Cap scan to 10000 rows to avoid loading huge files entirely into memory
    max_scan = min(ws.max_row or 1, 10000)
    for row in ws.iter_rows(max_row=max_scan):
        for cell in row:
            if cell.value is not None and query in str(cell.value).lower():
                matches.append({
                    "cell": cell.coordinate,
                    "value": _cell_value(cell),
                    "row": cell.row,
                    "column": cell.column,
                })
                if len(matches) >= args.limit:
                    break
        if len(matches) >= args.limit:
            break

    wb.close()
    print(json.dumps({
        "file": args.file,
        "sheet": ws.title,
        "query": args.query,
        "matches": len(matches),
        "results": matches,
    }, indent=2, default=str))


def cmd_to_csv(args):
    """Export a sheet to CSV."""
    _check_openpyxl()
    wb = _load_workbook(args.file, read_only=True)
    ws = _get_sheet(wb, args.sheet)

    output = args.output
    if not output:
        base = os.path.splitext(args.file)[0]
        sheet_suffix = f"_{ws.title}" if len(wb.sheetnames) > 1 else ""
        output = f"{base}{sheet_suffix}.csv"

    rows_written = 0
    max_export = 100000  # safety cap
    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows():
            writer.writerow([_safe_csv_value(_cell_value(c)) for c in row])
            rows_written += 1
            if rows_written >= max_export:
                break

    wb.close()
    print(json.dumps({
        "exported": output,
        "source": args.file,
        "sheet": ws.title,
        "rows": rows_written,
    }, indent=2))


def _detect_encoding(path: str) -> str:
    """Detect file encoding by checking BOM and common patterns."""
    with open(path, "rb") as f:
        head = f.read(4)
    if head[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    if head[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    # Try utf-8 first, fall back to latin-1 (never fails)
    try:
        with open(path, "r", encoding="utf-8") as f:
            f.read(4096)
        return "utf-8"
    except UnicodeDecodeError:
        pass
    for enc in ("gbk", "gb2312", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "utf-8"


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content length."""
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, (ws.max_column or 0) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                 max_row=min(ws.max_row or 1, 200)):
            cell = row[0]
            if cell.value is not None:
                # Count wide characters (CJK) as 2, others as 1
                text = str(cell.value)
                length = sum(2 if ord(c) > 0x4e00 else 1 for c in text)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _style_header(ws) -> None:
    """Apply professional header formatting: bold, colored background, borders."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    # Freeze header row
    ws.freeze_panes = "A2"


def cmd_from_csv(args):
    """Import CSV file(s) into a formatted XLSX workbook.

    Features: auto encoding detection, header styling, auto column widths,
    frozen header row, multi-file merge into separate sheets.
    """
    _check_openpyxl()
    import openpyxl

    csv_files = [args.csv_file]
    # Support multiple CSV files via extra_csvs
    if hasattr(args, "extra_csvs") and args.extra_csvs:
        csv_files.extend(args.extra_csvs)

    for f in csv_files:
        if not os.path.exists(f):
            sys.exit(f"CSV file not found: {f}")
    if os.path.exists(args.xlsx_file):
        sys.exit(f"Output file already exists: {args.xlsx_file}. Remove it first or use a different name.")

    # Parse sheet names
    sheet_names = []
    if hasattr(args, "sheet_names") and args.sheet_names:
        sheet_names = args.sheet_names
    elif args.sheet:
        sheet_names = [args.sheet]

    wb = openpyxl.Workbook()
    # Remove default sheet — we'll create named ones
    default_ws = wb.active

    total_rows = 0
    sheets_created = []
    max_import = 100000

    for i, csv_path in enumerate(csv_files):
        # Determine sheet name
        if i < len(sheet_names):
            name = sheet_names[i][:31]  # Excel limit
        else:
            name = os.path.splitext(os.path.basename(csv_path))[0][:31]

        if i == 0:
            default_ws.title = name
            ws = default_ws
        else:
            ws = wb.create_sheet(title=name)

        encoding = _detect_encoding(csv_path)
        rows_written = 0

        with open(csv_path, "r", encoding=encoding) as f:
            reader = csv.reader(f)
            for row in reader:
                if rows_written >= max_import:
                    break
                converted = []
                for val in row:
                    try:
                        if "." in val:
                            converted.append(float(val))
                        else:
                            converted.append(int(val))
                    except (ValueError, TypeError):
                        converted.append(_safe_value(val))
                ws.append(converted)
                rows_written += 1

        # Apply formatting if requested
        if not args.no_style and rows_written > 0:
            _style_header(ws)
            _auto_width(ws)

        total_rows += rows_written
        sheets_created.append({"name": name, "rows": rows_written, "encoding": encoding})

    wb.save(args.xlsx_file)
    wb.close()
    print(json.dumps({
        "created": args.xlsx_file,
        "sheets": sheets_created,
        "total_rows": total_rows,
    }, indent=2))


def cmd_add_row(args):
    """Append a row to a sheet."""
    _check_openpyxl()
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    try:
        data = json.loads(args.data)
        if not isinstance(data, list):
            sys.exit("--data must be a JSON array: [\"a\", \"b\", 42]")
    except json.JSONDecodeError as e:
        sys.exit(f"Invalid JSON in --data: {e}")

    ws.append([_safe_value(v) for v in data])
    new_row = ws.max_row

    wb.save(args.file)
    wb.close()
    print(json.dumps({
        "file": args.file,
        "sheet": ws.title,
        "appended_row": new_row,
        "data": data,
    }, indent=2, default=str))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        prog="excel_tool.py",
        description="Excel/XLSX tool for Hermes Agent — read, create, edit, search, convert",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # info
    p_info = sub.add_parser("info", help="Workbook metadata: sheets, dimensions")
    p_info.add_argument("file")

    # read
    p_read = sub.add_parser("read", help="Read cell data from a sheet")
    p_read.add_argument("file")
    p_read.add_argument("--sheet", help="Sheet name (default: active)")
    p_read.add_argument("--range", help="Cell range e.g. A1:C10")
    p_read.add_argument("--limit", type=int, default=100, help="Max rows (default: 100)")
    p_read.add_argument("--headers", action="store_true", help="Treat first row as headers")

    # create
    p_create = sub.add_parser("create", help="Create a new workbook")
    p_create.add_argument("file")
    p_create.add_argument("--sheet", help="Sheet name (default: Sheet1)")
    p_create.add_argument("--data", help='JSON array of rows: [[1,"a"],[2,"b"]]')
    p_create.add_argument("--from-json", help='JSON array of sheet defs: [{"name":"Sheet1","data":[[...]]}]')
    p_create.add_argument("--force", action="store_true", help="Overwrite existing file")

    # edit
    p_edit = sub.add_parser("edit", help="Edit a cell value")
    p_edit.add_argument("file")
    p_edit.add_argument("--cell", required=True, help="Cell reference e.g. A1, B3")
    p_edit.add_argument("--value", required=True, help="New value")
    p_edit.add_argument("--sheet", help="Sheet name (default: active)")

    # search
    p_search = sub.add_parser("search", help="Search for a value in cells")
    p_search.add_argument("file")
    p_search.add_argument("query")
    p_search.add_argument("--sheet", help="Sheet name (default: active)")
    p_search.add_argument("--limit", type=int, default=50, help="Max results (default: 50)")

    # to-csv
    p_csv = sub.add_parser("to-csv", help="Export sheet to CSV")
    p_csv.add_argument("file")
    p_csv.add_argument("--sheet", help="Sheet name (default: active)")
    p_csv.add_argument("--output", help="Output CSV path (default: auto)")

    # from-csv
    p_from = sub.add_parser("from-csv", help="Import CSV into formatted XLSX (auto-width, styled headers)")
    p_from.add_argument("csv_file", help="Primary CSV file")
    p_from.add_argument("xlsx_file", help="Output XLSX file")
    p_from.add_argument("extra_csvs", nargs="*", help="Additional CSV files (each becomes a sheet)")
    p_from.add_argument("--sheet", help="Sheet name for single file (default: filename)")
    p_from.add_argument("--sheet-names", nargs="*", help="Sheet names for multi-file (one per CSV)")
    p_from.add_argument("--no-style", action="store_true", help="Skip header styling and auto-width")

    # add-row
    p_add = sub.add_parser("add-row", help="Append a row to a sheet")
    p_add.add_argument("file")
    p_add.add_argument("--data", required=True, help='JSON array: ["a","b",42]')
    p_add.add_argument("--sheet", help="Sheet name (default: active)")

    args = parser.parse_args()
    dispatch = {
        "info": cmd_info,
        "read": cmd_read,
        "create": cmd_create,
        "edit": cmd_edit,
        "search": cmd_search,
        "to-csv": cmd_to_csv,
        "from-csv": cmd_from_csv,
        "add-row": cmd_add_row,
    }
    dispatch[args.command](args)


if __name__ == "__main__":
    main()
